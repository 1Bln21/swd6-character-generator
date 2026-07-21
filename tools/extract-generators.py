# -*- coding: utf-8 -*-
"""
=============================================================================
 Spieldaten für Droiden- und Schiffs-Generator nach gendata.js übertragen
=============================================================================

 Erzeugt gendata.js aus den Excel-Generatoren von Chance Gibboney:
   - "Droid Generator v1-3.xlsm"
   - "Ship Generator v1-1.xlsx"

 Aufruf:
     python tools/extract-generators.py "Pfad/Droid Generator.xlsm" "Pfad/Ship Generator.xlsx"

 Benötigt: pip install openpyxl

 Die Zeilen-/Spaltenbereiche entsprechen den o. g. Versionen. Nach dem
 Lauf die ausgegebene Zusammenfassung mit der vorherigen vergleichen.
=============================================================================
"""
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit('Bitte zuerst installieren:  pip install openpyxl')

DROID_SRC = sys.argv[1] if len(sys.argv) > 1 else 'Droid Generator v1-3.xlsm'
SHIP_SRC = sys.argv[2] if len(sys.argv) > 2 else 'Ship Generator v1-1.xlsx'
for p in (DROID_SRC, SHIP_SRC):
    if not os.path.isfile(p):
        sys.exit(f'Excel-Datei nicht gefunden: {p}')

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'gendata.js')


def s(ws, coord):
    x = ws[coord].value
    return '' if x is None else str(x).strip()


def num(ws, coord, default=0):
    x = ws[coord].value
    if x is None:
        return default
    try:
        return float(x) if isinstance(x, float) and x != int(x) else int(x)
    except (TypeError, ValueError):
        return default


def col_list(ws, col, r1, r2):
    out = []
    for r in range(r1, r2 + 1):
        v = s(ws, f'{col}{r}')
        if v:
            out.append(v)
    return out


# ========================================================= DROID GENERATOR
dwb = openpyxl.load_workbook(DROID_SRC, data_only=True)
mc = dwb['Model Creation']

degrees = []
for r in range(140, 145):                      # First..Fifth Degree
    degrees.append({
        'name': s(mc, f'B{r}'),
        # Kosten-Multiplikatoren (in %) je Attribut: DEX KNO MEC PER STR TEC
        'mult': [num(mc, f'{c}{r}') for c in ['C', 'D', 'E', 'F', 'G', 'H']],
    })

manufacturers = col_list(mc, 'K', 102, 146)
body_types = col_list(mc, 'B', 102, 109)
locomotion = col_list(mc, 'B', 112, 121)
droid_scales = col_list(mc, 'B', 124, 128)
matrix = col_list(mc, 'B', 131, 137)
db_skills = col_list(mc, 'AC', 101, 133)       # inkl. "None"
db_levels = []                                  # Datenbank-Bonus: Stufe -> Pips
for r in range(134, 140):
    db_levels.append({'label': s(mc, f'U{r}'), 'pips': num(mc, f'V{r}')})

mod_sheet = dwb['Modifications']
droid_mods = []
cat = ''
for r in range(4, 182):
    name = s(mod_sheet, f'B{r}')
    if not name:
        continue
    if mod_sheet[f'E{r}'].value is None:        # Kategoriezeile (kein Pip-Preis)
        cat = name
        continue
    droid_mods.append({
        'cat': cat,
        'name': name,
        'desc': s(mod_sheet, f'C{r}'),
        'pips': num(mod_sheet, f'E{r}'),        # Kosten bei der Erschaffung (Pips)
        'cp': num(mod_sheet, f'M{r}'),          # Kosten bei späterem Einbau (CP)
    })

DROID = {
    'degrees': degrees,
    'manufacturers': manufacturers,
    'bodyTypes': body_types,
    'locomotion': locomotion,
    'scales': droid_scales,
    'matrix': matrix,
    'dbSkills': db_skills,
    'dbLevels': db_levels,
    'mods': droid_mods,
    # Erschaffungs-Grundwerte laut Tabelle:
    'startDice': 25,      # 25D = 75 Punkte für Attribute + Skills + Modifikationen
    'attrMinPips': 3,     # Attribute starten bei 1D
    'attrMaxPips': 39,    # Maximum 13D
    'cpPerAttrPip': 20,   # CP-Kosten je Attributs-Pip (mal Degree-Multiplikator/100)
}

# ========================================================== SHIP GENERATOR
swb = openpyxl.load_workbook(SHIP_SRC, data_only=True)
sel = swb['Ship Selection']
md = swb['Modifications Data']
ws_mod = swb['Modifications']

ship_scales = col_list(sel, 'I', 134, 139)
pilot_skills = col_list(sel, 'I', 142, 151)
covers = col_list(sel, 'I', 154, 158)
hyper_mults = col_list(sel, 'I', 107, 131)     # None, x0.5 ... x25

crew = swb['Crew Skill Selection']
crew_skills = [s(crew, f'B{r}') for r in range(5, 54, 3) if s(crew, f'B{r}')]

wsel = swb['Weapons Selection']
fire_arcs = col_list(wsel, 'C', 111, 115)
gun_skills = col_list(wsel, 'J', 111, 114)
weapon_scales = col_list(wsel, 'C', 119, 124)

# Allgemeine Umbauten (fester Preis) – aus dem sichtbaren Modifications-Blatt
general_mods = []
for r in range(3, 17):
    name = s(ws_mod, f'B{r}')
    if not name:
        continue
    general_mods.append({
        'name': name,
        'desc': s(ws_mod, f'C{r}'),
        'cost': num(ws_mod, f'D{r}'),
        'weight': num(ws_mod, f'E{r}'),
        'avail': s(ws_mod, f'F{r}'),
    })

def pct_mods(label_col, diff_col, cost_col, mishap_col, r1, r2):
    """Prozentuale Umbauten: Stufe / Einbau-Schwierigkeit / Kosten (% vom Schiffswert) / Panne."""
    out = []
    for r in range(r1, r2 + 1):
        lab = s(md, f'{label_col}{r}')
        if not lab:
            continue
        out.append({'label': lab, 'diff': s(md, f'{diff_col}{r}'),
                    'costPct': num(md, f'{cost_col}{r}'), 'mishap': num(md, f'{mishap_col}{r}')})
    return out

drive_mods = pct_mods('Q', 'R', 'S', 'T', 3, 6)            # Space +1 .. +4
maneuver_mods = pct_mods('X', 'Y', 'Z', 'AA', 3, 7)        # +0D+1 .. +1D+2
hull_mods = pct_mods('AL', 'AM', 'AN', 'AO', 3, 7)
shield_mods = pct_mods('AS', 'AT', 'AU', 'AV', 3, 7)
weapon_dmg_mods = pct_mods('AZ', 'BA', 'BB', 'BC', 3, 7)
hyper_improve = pct_mods('AE', 'AF', 'AG', 'AH', 19, 22)   # Multiplikator-Verbesserung

repl_drives = []
for r in range(3, 8):
    name = s(md, f'BG{r}')
    if not name or name == 'Custom':
        continue
    repl_drives.append({
        'model': name, 'maker': s(md, f'BH{r}'), 'type': s(md, f'BI{r}'),
        'cost': num(md, f'BJ{r}'), 'weight': num(md, f'BK{r}'), 'avail': s(md, f'BL{r}'),
        'space': num(md, f'BM{r}'), 'atmCruise': num(md, f'BN{r}'), 'atmMax': num(md, f'BO{r}'),
        'special': s(md, f'BP{r}'),
    })

repl_hyper = []
for r in range(2, 8):
    model = s(md, f'BT{r}')
    if not model or model.startswith('Custom'):
        continue
    repl_hyper.append({
        'mult': s(md, f'BS{r}'), 'model': model, 'maker': s(md, f'BU{r}'),
        'cost': num(md, f'BV{r}'), 'weight': num(md, f'BW{r}'), 'avail': s(md, f'BX{r}'),
        'special': s(md, f'BY{r}'),
    })

shield_gens = []
for r in range(3, 7):
    rating = s(md, f'CB{r}')
    if not rating:
        continue
    shield_gens.append({'rating': rating, 'cost': num(md, f'CC{r}'),
                        'weight': num(md, f'CD{r}'), 'pips': num(md, f'CE{r}')})

cargo_mods = []
for r in range(3, 7):
    t = s(md, f'CH{r}')
    if not t:
        continue
    cargo_mods.append({'name': t, 'cost': num(md, f'CI{r}'), 'weight': num(md, f'CJ{r}')})

SHIP = {
    'scales': ship_scales,
    'pilotSkills': pilot_skills,
    'crewSkills': crew_skills,
    'covers': covers,
    'hyperMults': hyper_mults,
    'fireArcs': fire_arcs,
    'gunSkills': gun_skills,
    'weaponScales': weapon_scales,
    'generalMods': general_mods,
    'driveMods': drive_mods,
    'maneuverMods': maneuver_mods,
    'hullMods': hull_mods,
    'shieldMods': shield_mods,
    'weaponDmgMods': weapon_dmg_mods,
    'hyperImprove': hyper_improve,
    'replDrives': repl_drives,
    'replHyper': repl_hyper,
    'shieldGens': shield_gens,
    'cargoMods': cargo_mods,
    'maxWeapons': 6,
}

# ------------------------------------------------------------------ Ausgabe
with open(OUT, 'w', encoding='utf-8') as f:
    f.write('// Automatisch erzeugt aus "%s" und "%s"\n' % (os.path.basename(DROID_SRC), os.path.basename(SHIP_SRC)))
    f.write('// Quelle: Droiden-/Schiffs-Generatoren von Chance Gibboney (Star Wars D6, WEG)\n')
    f.write('// Nicht von Hand bearbeiten - stattdessen tools/extract-generators.py laufen lassen.\n')
    f.write('const DROID_DATA = ')
    json.dump(DROID, f, ensure_ascii=False, indent=1)
    f.write(';\nconst SHIP_DATA = ')
    json.dump(SHIP, f, ensure_ascii=False, indent=1)
    f.write(';\n')

print(f'Geschrieben: {OUT}')
print('--- Zusammenfassung ---')
print(f'  DROID: {len(degrees)} Degrees, {len(manufacturers)} Hersteller, {len(body_types)} Chassis,')
print(f'         {len(locomotion)} Fortbewegungen, {len(matrix)} Matrix-Stufen, {len(db_skills)} DB-Skills,')
print(f'         {len(droid_mods)} Modifikationen in {len(set(m["cat"] for m in droid_mods))} Kategorien')
print(f'  SHIP:  {len(general_mods)} allg. Umbauten, {len(drive_mods)}/{len(maneuver_mods)}/{len(hull_mods)}/{len(shield_mods)}/{len(weapon_dmg_mods)} %%-Umbauten,')
print(f'         {len(repl_drives)} Ersatz-Antriebe, {len(repl_hyper)} Ersatz-Hyperantriebe, {len(shield_gens)} Schildgeneratoren,')
print(f'         {len(cargo_mods)} Frachtumbauten, {len(hyper_mults)} Hyper-Multiplikatoren, {len(ship_scales)} Größenklassen')
