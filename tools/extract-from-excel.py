# -*- coding: utf-8 -*-
"""
=============================================================================
 Carry the game data from the original spreadsheet over into data.js
=============================================================================

 Generates data.js from Chance Gibboney's Excel character generator
 ("Character Generator v2-5.xlsx" or newer).

 Usage:
     python tools/extract-from-excel.py "path/to/Character Generator.xlsx"

 With no argument it looks for "Character Generator v2-5.xlsx" in the
 current folder. The result lands as data.js beside the project folder.

 Requires: pip install openpyxl

 -----------------------------------------------------------------------
 Important with a NEW version of the workbook:
 The row and column ranges below refer to the v2.5 layout. If that layout
 changes, the ranges have to be adjusted. After a run the script prints a
 summary (number of species, skills, Force powers ...) - compare those
 figures against the previous version and a range that has slipped shows
 up at once.
 -----------------------------------------------------------------------
=============================================================================
"""
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit('Bitte zuerst installieren:  pip install openpyxl')

SRC = sys.argv[1] if len(sys.argv) > 1 else 'Character Generator v2-5.xlsx'
if not os.path.isfile(SRC):
    sys.exit(f'Excel-Datei nicht gefunden: {SRC}')

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data.js')
wb = openpyxl.load_workbook(SRC, data_only=True)

ATTR_KEYS = ['dex', 'kno', 'mec', 'per', 'str', 'tec']


def v(ws, coord):
    """The cell value, but 0/'0'/'None' all count as empty."""
    x = ws[coord].value
    if x is None:
        return None
    if isinstance(x, str):
        x = x.strip()
        if x in ('0', 'None', ''):
            return None
    if isinstance(x, (int, float)) and x == 0:
        return None
    return x


def num(ws, coord, default=0):
    x = ws[coord].value
    if x is None:
        return default
    try:
        return float(x) if isinstance(x, float) and x != int(x) else int(x)
    except (TypeError, ValueError):
        return default


def s(ws, coord):
    x = ws[coord].value
    return '' if x is None else str(x).strip()


def dec(val):
    """'1,5' -> 1.5 (the sheet sometimes uses a comma as decimal point)."""
    if isinstance(val, str):
        try:
            return float(val.replace(',', '.'))
        except ValueError:
            return 0
    return val if val is not None else 0


# ---------------------------------------------------------------- species
r1 = wb['Race Information 1']      # attribute limits, move, starting dice
r2 = wb['Race Information 2']      # abilities, story factors, bonus skills

ATTR_TAG = {'DEX': 'dex', 'KNOW': 'kno', 'MECH': 'mec',
            'PERC': 'per', 'STR': 'str', 'TECH': 'tec'}


def species_row(row, name_override=None):
    sp = {
        'name': name_override or s(r1, f'B{row}'),
        'min':  [num(r1, f'{c}{row}') for c in ['C', 'E', 'G', 'I', 'K', 'M']],
        'max':  [num(r1, f'{c}{row}') for c in ['D', 'F', 'H', 'J', 'L', 'N']],
        'move': num(r1, f'O{row}'),
        'free': num(r1, f'P{row}'),
        'offset': num(r1, f'Q{row}'),
        'hMin': dec(r1[f'R{row}'].value),
        'hMax': dec(r1[f'S{row}'].value),
        'planet': s(r1, f'T{row}'),
        'page': s(r1, f'U{row}'),
        'abilities': [], 'story': [], 'skillImprove': [], 'bonusSkills': [],
        'armorP': num(r2, f'I{row}'), 'armorE': num(r2, f'J{row}'),
    }
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:          # special abilities
        a = v(r2, f'{col}{row}')
        if a:
            sp['abilities'].append(str(a).strip())
    for col in ['O', 'P', 'Q', 'R']:                     # story factors
        a = v(r2, f'{col}{row}')
        if a:
            sp['story'].append(str(a).strip())
    for col in ['K', 'L', 'M', 'N']:                     # typical skills
        a = v(r2, f'{col}{row}')
        if a:
            sp['skillImprove'].append(str(a).strip())
    for col in ['S', 'T', 'U']:                          # bonus skills
        a = v(r2, f'{col}{row}')
        if not a:
            continue
        m = re.match(r'^(.*?)\s*\((DEX|KNOW|MECH|PERC|STR|TECH)\)\s*$', str(a).strip())
        if m:
            sp['bonusSkills'].append({'name': m.group(1).strip(),
                                      'attr': ATTR_TAG[m.group(2)]})
        else:
            sp['bonusSkills'].append({'name': str(a).strip(), 'attr': 'kno'})
    return sp


species = [species_row(i) for i in range(2, 62)]          # 60 species
near_humans = [species_row(i) for i in range(64, 73)]      # 9 near-human variants
trianii = {'Female': species_row(76, 'Trianii (Female)'),
           'Male':   species_row(77, 'Trianii (Male)')}

# ------------------------------------------------------------------ skills
def skill_list(sheet):
    ws = wb[sheet]
    return [str(v(ws, f'M{r}')).strip() for r in range(10, 58) if v(ws, f'M{r}')]


skills = {
    'dex': skill_list('Skill Selection - Dexterity'),
    'kno': skill_list('Skill Selection - Knowledge'),
    'mec': skill_list('Skill Selection - Mechanical'),
    'per': skill_list('Skill Selection - Perception'),
    'str': skill_list('Skill Selection - Strength'),
    'tec': skill_list('Skill Selection - Technical'),
}

# ----------------------------------------------------------- Force powers
fp = wb['Force Powers']
fc = wb['Force Powers Calculations']

meta = {}
for r in range(4, 96):
    nm = fc[f'A{r}'].value
    if nm and str(nm).strip() not in ('Control', 'Sense', 'Alter'):
        meta[str(nm).strip()] = {'diff': s(fc, f'B{r}'),
                                 'kept': s(fc, f'C{r}'),
                                 'dark': s(fc, f'D{r}')}

CAT_RANGES = [
    ('Control', 5, 24), ('Sense', 26, 41), ('Alter', 43, 47),
    ('Control & Sense', 49, 54), ('Control & Alter', 56, 71),
    ('Sense & Alter', 73, 75), ('Control, Sense & Alter', 77, 91),
    ('Special', 93, 93),
]
powers = []
for cat, a, b in CAT_RANGES:
    for r in range(a, b + 1):
        nm = fp[f'E{r}'].value
        if not nm:
            continue
        nm = str(nm).strip()
        m = meta.get(nm, {})
        powers.append({'name': nm, 'cat': cat,
                       'prereq': s(fp, f'I{r}'), 'page': s(fp, f'R{r}'),
                       'diff': m.get('diff', ''), 'kept': m.get('kept', ''),
                       'dark': m.get('dark', '')})

# -------------------------------------------------------------- equipment
eq = wb['Equipment']
CATS = {'Communication', 'General', 'Medical', 'Restraining Devices',
        'Special Tools', 'Surveillance', 'Transport', 'Travel Aids',
        'Other Equipment'}
equipment, cat = [], ''
for r in range(2, 84):
    a = eq[f'A{r}'].value
    if a is None:
        continue
    a = str(a).strip()
    if a in CATS:
        if a == 'Other Equipment':
            break
        cat = a
        continue
    equipment.append({'cat': cat, 'name': a, 'cost': num(eq, f'B{r}'),
                      'avail': s(eq, f'C{r}'), 'note': s(eq, f'D{r}')})

# ------------------------------------------------------------------ armor
df = wb['Defense']
armor = []
for r in range(151, 191):
    nm = df[f'C{r}'].value
    if not nm:
        continue
    abilities = [str(v(df, f'{c}{r}')).strip()
                 for c in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q'] if v(df, f'{c}{r}')]
    armor.append({'name': str(nm).strip(), 'cost': num(df, f'D{r}'),
                  'avail': s(df, f'E{r}'), 'phys': num(df, f'F{r}'),
                  'energy': num(df, f'G{r}'), 'loc': s(df, f'H{r}'),
                  'dexPen': num(df, f'I{r}'), 'abilities': abilities})

# --------------------------------------------------------------- melee
ml = wb['Melee Attack']
melee = []
for r in range(152, 168):
    nm = ml[f'C{r}'].value
    if not nm:
        continue
    nm = str(nm).strip()
    if nm.startswith('Lightsaber: Custom'):     # comes from the workshop
        continue
    melee.append({'name': nm, 'cost': num(ml, f'D{r}'), 'avail': s(ml, f'E{r}'),
                  'dmg': num(ml, f'F{r}'), 'maxDmg': num(ml, f'G{r}'),
                  'diff': s(ml, f'H{r}'), 'ability': s(ml, f'I{r}'),
                  'color': s(ml, f'J{r}')})

# -------------------------------------------------------------- ranged
rg = wb['Ranged Attack']
ranged = []
for r in range(148, 181):
    nm = rg[f'C{r}'].value
    if not nm:
        continue
    ranged.append({'name': str(nm).strip(), 'cost': num(rg, f'D{r}'),
                   'avail': s(rg, f'E{r}'), 'dmg': num(rg, f'F{r}'),
                   'close': s(rg, f'G{r}'), 'short': s(rg, f'H{r}'),
                   'medium': s(rg, f'I{r}'), 'long': s(rg, f'J{r}'),
                   'rof': s(rg, f'K{r}'), 'ammo': s(rg, f'L{r}'),
                   'ability': s(rg, f'M{r}'), 'skill': s(rg, f'N{r}') or 'Blaster'})

# ------------------------------------------------------------ explosives
ex = wb['Explosives']
explosives = []
for r in range(132, 139):
    nm = ex[f'C{r}'].value
    if not nm:
        continue
    explosives.append({'name': str(nm).strip(), 'cost': num(ex, f'D{r}'),
                       'avail': s(ex, f'E{r}'),
                       'dmg':    [s(ex, f'{c}{r}') for c in ['F', 'G', 'H', 'I']],
                       'ranges': [s(ex, f'{c}{r}') for c in ['J', 'K', 'L', 'M']],
                       'radius': [s(ex, f'{c}{r}') for c in ['N', 'O', 'P', 'Q']],
                       'ability': s(ex, f'R{r}')})

# ------------------------------------------------------------ lightsabers
ls = wb['Custom Lightsaber']
primary = [{'name': str(ls[f'C{r}'].value).strip(), 'color': s(ls, f'D{r}'),
            'dmg': num(ls, f'E{r}'), 'ability': s(ls, f'F{r}')}
           for r in range(151, 167)
           if ls[f'C{r}'].value and str(ls[f'C{r}'].value).strip() != 'None']
secondary = [{'name': str(ls[f'K{r}'].value).strip(), 'color': s(ls, f'L{r}'),
              'mod': num(ls, f'M{r}'), 'ability': s(ls, f'N{r}')}
             for r in range(152, 180) if ls[f'K{r}'].value]
mods = [{'name': str(ls[f'Y{r}'].value).strip(), 'cost': num(ls, f'Z{r}'),
         'ability': s(ls, f'AA{r}')}
        for r in range(152, 168) if ls[f'Y{r}'].value]
saber_colors = [str(ls[f'AK{r}'].value).strip()
                for r in range(151, 158) if ls[f'AK{r}'].value]

# ------------------------------------------------------- planets, gender
pi = wb['Personal Information']
planets = sorted({str(pi[f'P{r}'].value).strip()
                  for r in range(136, 229) if pi[f'P{r}'].value})
genders = [str(pi[f'H{r}'].value).strip()
           for r in range(136, 141) if pi[f'H{r}'].value]

# ------------------------------------------------------------------- output
data = {
    'species': species, 'nearHumans': near_humans, 'trianii': trianii,
    'skills': skills, 'powers': powers, 'equipment': equipment, 'armor': armor,
    'melee': melee, 'ranged': ranged, 'explosives': explosives,
    'saber': {'primary': primary, 'secondary': secondary,
              'mods': mods, 'colors': saber_colors},
    'planets': planets, 'genders': genders,
}

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('// Generated from "%s"\n' % os.path.basename(SRC))
    f.write('// Source: Chance Gibboney\'s Excel character generator\n')
    f.write('// (Star Wars D6, 2nd Edition - West End Games)\n')
    f.write('// Do not edit by hand - run tools/extract-from-excel.py instead.\n')
    f.write('const DATA = ')
    json.dump(data, f, ensure_ascii=False, indent=1)
    f.write(';\n')

print(f'Geschrieben: {OUT}')
print('--- Zusammenfassung (mit der vorherigen Fassung vergleichen!) ---')
print(f'  Spezies:        {len(species)}  + {len(near_humans)} Near-Human-Varianten')
print(f'  Fertigkeiten:   ' + ', '.join(f'{k}={len(x)}' for k, x in skills.items()))
print(f'  Machtkräfte:    {len(powers)}')
print(f'  Ausrüstung:     {len(equipment)}')
print(f'  Rüstungen:      {len(armor)}')
print(f'  Nahkampf:       {len(melee)}')
print(f'  Fernkampf:      {len(ranged)}')
print(f'  Sprengstoffe:   {len(explosives)}')
print(f'  Lichtschwert:   {len(primary)} Primär-, {len(secondary)} Sekundärkristalle, '
      f'{len(mods)} Modifikationen, {len(saber_colors)} Farben')
print(f'  Planeten:       {len(planets)}')
print(f'  Geschlechter:   {len(genders)}')
